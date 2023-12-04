from contextlib import suppress
from csv import Error, Sniffer, reader
from ctypes import WinDLL, c_uint
from itertools import chain, zip_longest
from logging import DEBUG, INFO, basicConfig, debug, error, info, root, warning
from os import name, system
from pathlib import Path
from re import sub
from string import ascii_lowercase as ascii
from traceback import print_exc
from typing import Dict, Final, Iterable, Optional

from click import Context
from click import Path as click_Path
from click import argument, command, echo, option, pass_context
from openpyxl import load_workbook
from xlrd import open_workbook
from xlrd.sheet import Sheet

invalid_threshold: Final = 0.5
valid_threshold: Final = 0.9
excel_sigs: Final = [
    ('xlsx', b'\x50\x4B\x05\x06', 2, -22, 4),
    ('xls', b'\x09\x08\x10\x00\x00\x06\x05\x00', 0, 512, 8),  #  Excel
    ('xls', b'\x09\x08\x10\x00\x00\x06\x05\x00', 0, 1536, 8),  # Calc
    ('xls', b'\x09\x08\x10\x00\x00\x06\x05\x00', 0, 2048, 8),  # Excel > Calc
]


@command('joiner', context_settings=dict(ignore_unknown_options=True))
@argument(
    'input',
    nargs=-1,
    type=click_Path(exists=True, path_type=Path),
)
@option(
    '-o',
    '--output',
    help='The path to output data to.',
    default='numbers.txt',
    type=click_Path(dir_okay=False, path_type=Path),
    required=True,
)
@option(
    '-c',
    '--column',
    help='The column (e.g. A, B, AB) to process from table.',
    callback=lambda ctx, param, value: value.upper() if value else None,
)
@option(
    '--all/--no-all',
    help='If all suggested columns should be used from table (>90% valid).',
    is_flag=True,
    default=False,
)
@option(
    '--sort/--no-sort',
    help='Whether export data should be sorted.',
    is_flag=True,
    default=True,
)
@option(
    '-l',
    '--logging',
    help='The logging level used to print information.',
    default='INFO',
    callback=lambda ctx, param, value: value.upper() if value else None,
)
@pass_context
def cli(
    context: Context,
    /,
    input: Iterable[Path],
    output: Path,
    column: Optional[str] = None,
    *,
    all: Optional[bool] = None,
    sort: Optional[bool] = None,
    logging: Optional[str] = None,
) -> None:
    basicConfig(level=(logging or '').upper(), force=True)
    input = list(
        chain.from_iterable(
            _.rglob('*') if _.is_dir() else (_,) for _ in input
        )
    )
    if not input:
        return echo(cli.get_help(context))
    data = {}
    for file in input:
        # if file.name.startswith('~$'):
        #     continue
        debug('Reading `%s`...', file)

        try:
            if excel_type := get_excel_type(file):
                if excel_type == 'xls':
                    new_data = read_xls(file, column=column, is_all=all)
                else:
                    new_data = read_xlsx(file, column=column, is_all=all)
            else:
                new_data = read_csv(file, column=column, is_all=all)
        except BaseException as _:
            if root.level == DEBUG:
                print_exc()
            else:
                error(_)
            continue
        if not new_data:
            info('Got nothing from `%s`!', file)
        else:
            info(
                'Got `%s` phone%s from `%s` (`%s` duplicate%s)!',
                new := sum(_ not in data for _ in new_data),
                '' if new == 1 else 's',
                file,
                duplicates := sum(_ in data for _ in new_data),
                '' if duplicates == 1 else 's',
            )
        data |= new_data

    if not data:
        return info('Nothing to export!')
    info(
        'Exporting `%s` phone%s to `%s`!',
        len(data),
        '' if len(data) == 1 else 's',
        output,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, 'w') as fp:
        for phone in sorted(data) if sort else data:
            fp.write(phone + '\n')


def format(line: object, /) -> None:
    if (line_len := len(line := sub('\D', '', str(line)))) == 11:
        return line
    elif line_len == 10:
        return '7' + line[-10:]
    else:
        return None


def get_column_index(column: str, /) -> int:
    total = None
    for index, letter in enumerate(column.lower()):
        with suppress(ValueError):
            total = (total or 0) + index * len(ascii) + ascii.index(letter)
    return total


def get_excel_type(file: str, /) -> Optional[str]:
    for sigType, sig, whence, offset, size in excel_sigs:
        with open(file, 'rb') as fp:
            fp.seek(offset, whence)
            if fp.read(size) == sig:
                return sigType
    return None


def read_rows(
    rows: Iterable[Iterable[object]],
    /,
    column: Optional[str] = None,
    is_all: Optional[bool] = None,
) -> Dict[str, None]:
    cols = enumerate(zip_longest(*rows, fillvalue=None))
    if column:
        column_index = get_column_index(column)
        for index, col in cols:
            if index == column_index:
                return dict.fromkeys(p for c in col if (p := format(c)))
        return {}

    data, col_data, col_stat = {}, {}, {}
    for index, col in cols:
        _col_data = [p for c in col if (p := format(c))]
        col_data[index] = dict.fromkeys(_col_data)
        col_stat[index] = len(_col_data) / len(col)
        if col_stat[index] > valid_threshold:
            data |= col_data[index]
            if not is_all:
                break
    if not data:
        index = sorted(col_stat, key=col_stat.get)[-1]
        if col_stat[index] > invalid_threshold:
            return col_data[index]
    return data


def read_csv(
    file: Path,
    /,
    column: Optional[str] = None,
    is_all: Optional[bool] = None,
) -> Dict[str, None]:
    with open(file, encoding='utf-8', newline='') as fp:
        try:
            dialect = Sniffer().sniff(
                fp.readline(),
                delimiters=[',', ';', '|'],
            )
        except Error:
            dialect = None
        fp.seek(0)
        return read_rows(reader(fp, dialect=dialect), column, is_all=is_all)


def read_xls(
    file: Path,
    /,
    column: Optional[str] = None,
    is_all: Optional[bool] = None,
) -> Dict[str, None]:
    data = {}
    sheet: Sheet
    for sheet in open_workbook(file, on_demand=False).sheets():
        rows = (sheet.row_values(index) for index in range(sheet.nrows))
        data |= read_rows(rows, column, is_all=is_all)
    return data


def read_xlsx(
    file: Path,
    /,
    column: Optional[str] = None,
    is_all: Optional[bool] = None,
) -> Dict[str, None]:
    data = {}
    for sheet in load_workbook(filename=file, read_only=True).worksheets:
        rows = sheet.iter_rows(values_only=True)
        data |= read_rows(rows, column, is_all=is_all)
    return data


if __name__ == '__main__':
    try:
        cli()
    finally:
        if name == 'nt' and root.level <= INFO:
            kernel32 = WinDLL('kernel32', use_last_error=True)
            process_array = (c_uint * 1)()
            if kernel32.GetConsoleProcessList(process_array, 1) == 2:
                system('pause')
